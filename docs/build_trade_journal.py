"""FX Trade Journal (xlsx) を生成する。

3 シート構成:
  1. ダッシュボード — 自動集計の KPI / 手法別・ペア別パフォーマンス
  2. 取引履歴       — 1 トレード 1 行で記入。自動計算カラムを含む
  3. 設定           — 開始残高 / リスク% / pip 価値テーブル

Google Sheets に直接アップロード可能。
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule

# ============================================================
# テーマカラー (Google Sheets でも見やすいよう、ヘッダーは濃色、データは淡色)
# ============================================================
CLR_HEADER_BG    = "1F1B4E"     # 濃紫
CLR_HEADER_FG    = "F5ECD7"     # アイボリー
CLR_GOLD         = "E9C46A"     # 金
CLR_GOLD_BG      = "FFF7E0"     # 薄金
CLR_GREEN        = "16A34A"     # 緑
CLR_GREEN_BG     = "DCFCE7"     # 薄緑
CLR_RED          = "DC2626"     # 赤
CLR_RED_BG       = "FEE2E2"     # 薄赤
CLR_ROW_ALT_BG   = "F8F4FF"     # 淡紫 (シマシマ用)
CLR_PANEL_BG     = "F3F0FF"     # パネル背景
CLR_BORDER       = "C7C0E0"     # 罫線

# ============================================================
# 共通スタイル
# ============================================================
FONT_BASE      = Font(name="Noto Sans JP", size=10)
FONT_BASE_BOLD = Font(name="Noto Sans JP", size=10, bold=True)
FONT_HEADER    = Font(name="Noto Sans JP", size=10, bold=True, color=CLR_HEADER_FG)
FONT_TITLE     = Font(name="Noto Sans JP", size=18, bold=True, color=CLR_HEADER_BG)
FONT_KPI       = Font(name="Noto Sans JP", size=16, bold=True, color=CLR_HEADER_BG)
FONT_KPI_GOLD  = Font(name="Noto Sans JP", size=16, bold=True, color="B8860B")
FONT_LABEL     = Font(name="Noto Sans JP", size=9, color="6B5B95")
FONT_NOTE      = Font(name="Noto Sans JP", size=9, italic=True, color="6B5B95")

THIN  = Side(style="thin", color=CLR_BORDER)
THICK = Side(style="medium", color=CLR_HEADER_BG)
BORDER_ALL  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOX  = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

FILL_HEADER = PatternFill("solid", fgColor=CLR_HEADER_BG)
FILL_GOLD   = PatternFill("solid", fgColor=CLR_GOLD_BG)
FILL_PANEL  = PatternFill("solid", fgColor=CLR_PANEL_BG)
FILL_GREEN  = PatternFill("solid", fgColor=CLR_GREEN_BG)
FILL_RED    = PatternFill("solid", fgColor=CLR_RED_BG)
FILL_ALT    = PatternFill("solid", fgColor=CLR_ROW_ALT_BG)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center", indent=1)
RIGHT  = Alignment(horizontal="right", vertical="center")

# ============================================================
# 各シートで使う定数
# ============================================================
PAIRS = [
    "USD/JPY", "EUR/USD", "GBP/USD", "AUD/USD", "NZD/USD",
    "USD/CAD", "USD/CHF",
    "EUR/JPY", "GBP/JPY", "AUD/JPY", "NZD/JPY",
    "CAD/JPY", "CHF/JPY",
    "EUR/GBP",
]
METHODS  = ["TRIPLE", "ORZ", "PDHL", "BOTH", "CLAUDE"]
DIRECTIONS = ["Long", "Short"]
REASONS  = ["TP", "SL", "手動", "強制", "建値撤退"]

MAX_TRADES = 500   # 取引履歴シートの行数上限


def make_workbook(path: str):
    wb = Workbook()
    # デフォルトの最初のシートを削除
    wb.remove(wb.active)

    # === シート作成 (順番が表示順) ===
    ws_dash = wb.create_sheet("ダッシュボード")
    ws_log  = wb.create_sheet("取引履歴")
    ws_cfg  = wb.create_sheet("設定")

    build_settings(ws_cfg)
    build_trade_log(ws_log)
    build_dashboard(ws_dash)

    # ワークブック全体のフォント
    for ws in wb:
        ws.sheet_view.showGridLines = False  # Google Sheets でもクリーンに

    wb.save(path)
    print(f"[ok] saved: {path}")


# ============================================================
# シート 1: ダッシュボード
# ============================================================
def build_dashboard(ws):
    # 列幅
    widths = [2, 18, 14, 14, 14, 14, 14, 14, 4]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ===== タイトル =====
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = "📊  FX Trade Journal · Dashboard"
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:H3")
    sub = ws["B3"]
    sub.value = "毎日 5 分、トレードしたら『取引履歴』シートに 1 行追加するだけで自動集計されます。"
    sub.font = FONT_NOTE
    sub.alignment = Alignment(horizontal="left", indent=1)

    # ===== KPI カード (B5:H8) =====
    # 6 つのカードを横並びに
    kpis = [
        ("総トレード",      "=COUNT(取引履歴!P:P)",                         "0",     None),
        ("勝率",           "=IFERROR(COUNTIF(取引履歴!P:P,\">0\")/COUNT(取引履歴!P:P),0)", "0.0%",  None),
        ("累計 R",         "=SUM(取引履歴!P:P)",                            "+0.00;-0.00;-",  "RR"),
        ("期待値 (R)",     "=IFERROR(AVERAGE(取引履歴!P:P),0)",            "+0.000;-0.000;-",  "per trade"),
        ("Profit Factor", '=IFERROR(SUMIF(取引履歴!P:P,">0")/-SUMIF(取引履歴!P:P,"<0"),0)', "0.00",  None),
        ("累計損益 (¥)",   "=SUM(取引履歴!S:S)",                            "¥#,##0;[Red]-¥#,##0;-", None),
    ]
    start_col = 2  # B
    for idx, (label, formula, fmt, sub_lbl) in enumerate(kpis):
        col = start_col + idx
        col_letter = get_column_letter(col)
        # ラベル
        ws[f"{col_letter}5"] = label
        ws[f"{col_letter}5"].font = FONT_LABEL
        ws[f"{col_letter}5"].alignment = CENTER
        ws[f"{col_letter}5"].fill = FILL_PANEL
        # 値
        ws[f"{col_letter}6"] = formula
        ws[f"{col_letter}6"].font = FONT_KPI_GOLD if idx == 2 else FONT_KPI
        ws[f"{col_letter}6"].alignment = CENTER
        ws[f"{col_letter}6"].fill = FILL_PANEL
        ws[f"{col_letter}6"].number_format = fmt
        # サブラベル
        ws[f"{col_letter}7"] = sub_lbl or ""
        ws[f"{col_letter}7"].font = FONT_LABEL
        ws[f"{col_letter}7"].alignment = CENTER
        ws[f"{col_letter}7"].fill = FILL_PANEL
        # 罫線
        for r in (5, 6, 7):
            ws[f"{col_letter}{r}"].border = BORDER_ALL

    # 高さ調整
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 34
    ws.row_dimensions[7].height = 16

    # ===== セクション: リスク管理 =====
    section_header(ws, "B10:D10", "🛡  リスク管理 / 連敗連勝")

    risk_rows = [
        ("最大連勝",       '=IF(COUNT(取引履歴!P:P)=0,0,MAX(取引履歴!U:U))',          "0"),
        ("最大連敗",       '=IF(COUNT(取引履歴!P:P)=0,0,MIN(取引履歴!U:U))',          "0"),
        ("最大ドローダウン", '=IF(COUNT(取引履歴!P:P)=0,0,MAX(取引履歴!Q:Q)-MIN(取引履歴!Q:Q))', "+0.00;-0.00;-"),
        ("勝ち平均 R",     '=IFERROR(AVERAGEIF(取引履歴!P:P,">0"),0)',                "+0.000;-0.000;-"),
        ("負け平均 R",     '=IFERROR(AVERAGEIF(取引履歴!P:P,"<0"),0)',                "+0.000;-0.000;-"),
        ("現在の残高",     '=設定!B2+SUM(取引履歴!S:S)',                              "¥#,##0"),
    ]
    for i, (label, formula, fmt) in enumerate(risk_rows):
        row = 11 + i
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = FONT_BASE
        ws[f"B{row}"].alignment = LEFT
        ws[f"B{row}"].fill = FILL_PANEL
        ws[f"B{row}"].border = BORDER_ALL

        ws.merge_cells(f"C{row}:D{row}")
        ws[f"C{row}"] = formula
        ws[f"C{row}"].font = FONT_BASE_BOLD
        ws[f"C{row}"].alignment = RIGHT
        ws[f"C{row}"].number_format = fmt
        ws[f"C{row}"].fill = FILL_PANEL
        for col in "CD":
            ws[f"{col}{row}"].border = BORDER_ALL

    # ===== セクション: 手法別パフォーマンス =====
    section_header(ws, "E10:H10", "🎯  手法別パフォーマンス")

    headers = ["手法", "件数", "勝率", "EV (R)"]
    for i, h in enumerate(headers):
        ws.cell(row=11, column=5+i, value=h)
        cell = ws.cell(row=11, column=5+i)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = BORDER_ALL

    for i, m in enumerate(METHODS):
        row = 12 + i
        ws.cell(row=row, column=5, value=m).font = FONT_BASE_BOLD
        ws.cell(row=row, column=5).alignment = LEFT
        ws.cell(row=row, column=5).fill = FILL_GOLD if m == "TRIPLE" else FILL_PANEL
        # 件数
        ws.cell(row=row, column=6, value=f'=COUNTIF(取引履歴!D:D,"{m}")').number_format = "0"
        # 勝率
        ws.cell(row=row, column=7,
                value=f'=IFERROR(COUNTIFS(取引履歴!D:D,"{m}",取引履歴!P:P,">0")/COUNTIF(取引履歴!D:D,"{m}"),0)'
                ).number_format = "0.0%"
        # EV (R)
        ws.cell(row=row, column=8,
                value=f'=IFERROR(SUMIF(取引履歴!D:D,"{m}",取引履歴!P:P)/COUNTIF(取引履歴!D:D,"{m}"),0)'
                ).number_format = "+0.000;-0.000;-"

        for c in range(5, 9):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER_ALL
            cell.alignment = RIGHT if c > 5 else LEFT
            if cell.font.size is None:
                cell.font = FONT_BASE

    # ===== セクション: ペア別パフォーマンス (上位 15) =====
    section_header(ws, "B18:H18", "💱  通貨ペア別パフォーマンス")
    headers2 = ["通貨ペア", "件数", "勝率", "累計 R", "平均 R", "最終トレード日付"]
    for i, h in enumerate(headers2):
        cell = ws.cell(row=19, column=2+i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = BORDER_ALL

    for i, p in enumerate(PAIRS):
        row = 20 + i
        ws.cell(row=row, column=2, value=p).font = FONT_BASE_BOLD
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=3, value=f'=COUNTIF(取引履歴!C:C,"{p}")').number_format = "0"
        ws.cell(row=row, column=4,
                value=f'=IFERROR(COUNTIFS(取引履歴!C:C,"{p}",取引履歴!P:P,">0")/COUNTIF(取引履歴!C:C,"{p}"),0)'
                ).number_format = "0.0%"
        ws.cell(row=row, column=5,
                value=f'=SUMIF(取引履歴!C:C,"{p}",取引履歴!P:P)'
                ).number_format = "+0.00;-0.00;-"
        ws.cell(row=row, column=6,
                value=f'=IFERROR(SUMIF(取引履歴!C:C,"{p}",取引履歴!P:P)/COUNTIF(取引履歴!C:C,"{p}"),0)'
                ).number_format = "+0.000;-0.000;-"
        ws.cell(row=row, column=7,
                value=f'=IFERROR(MAXIFS(取引履歴!B:B,取引履歴!C:C,"{p}"),"")'
                ).number_format = "yyyy/mm/dd"

        # 行スタイル
        for c in range(2, 8):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER_ALL
            if c == 2:
                cell.alignment = LEFT
            else:
                cell.alignment = RIGHT
            if cell.font.size is None:
                cell.font = FONT_BASE
            if i % 2 == 1:
                cell.fill = FILL_ALT

    # 累計 R に Color Scale (赤→白→緑)
    ws.conditional_formatting.add(
        f"E20:E{19+len(PAIRS)}",
        ColorScaleRule(start_type="min", start_color=CLR_RED_BG,
                       mid_type="num", mid_value=0, mid_color="FFFFFF",
                       end_type="max", end_color=CLR_GREEN_BG)
    )

    # ===== ガイドメッセージ =====
    guide_row = 20 + len(PAIRS) + 2
    ws.merge_cells(f"B{guide_row}:H{guide_row}")
    cell = ws.cell(row=guide_row, column=2)
    cell.value = "💡  使い方:  右の「取引履歴」シートに、トレードした日付・ペア・手法などを入力するだけ。すべての集計はこのページで自動更新されます。"
    cell.font = FONT_NOTE
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    cell.fill = FILL_GOLD
    cell.border = BORDER_BOX
    ws.row_dimensions[guide_row].height = 28


def section_header(ws, range_str, text):
    """セクション見出し用の merged cell を作る。"""
    ws.merge_cells(range_str)
    first = range_str.split(":")[0]
    cell = ws[first]
    cell.value = text
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    row = int("".join(filter(str.isdigit, first)))
    ws.row_dimensions[row].height = 22


# ============================================================
# シート 2: 取引履歴
# ============================================================
def build_trade_log(ws):
    headers = [
        ("#",          5,  "0"),
        ("日付",       12, "yyyy/mm/dd"),
        ("通貨ペア",   10, "@"),
        ("手法",       10, "@"),
        ("方向",       8,  "@"),
        ("スコア",     8,  "0"),
        ("Entry時刻", 10, "hh:mm"),
        ("Entry価格", 12, "0.00000"),
        ("SL価格",    12, "0.00000"),
        ("TP価格",    12, "0.00000"),
        ("ロット",     8,  "0.00"),
        ("決済価格",  12, "0.00000"),
        ("決済理由",   10, "@"),
        ("SL pips",   10, "0.0"),
        ("計画RR",    9,  "0.00"),
        ("損益 pips", 10, "+0.0;-0.0;-"),
        ("R-mult",   10, "+0.00;-0.00;-"),
        ("累計 R",   10, "+0.00;-0.00;-"),
        ("損益 ¥",   12, "¥#,##0;[Red]-¥#,##0;-"),
        ("メモ",      30, "@"),
    ]
    for i, (h, w, _fmt) in enumerate(headers, 1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = w
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[1].height = 28

    # データ範囲 (2 行目 〜 MAX_TRADES+1 行目)
    for r in range(2, MAX_TRADES + 2):
        for i, (_h, _w, fmt) in enumerate(headers, 1):
            cell = ws.cell(row=r, column=i)
            cell.font = FONT_BASE
            cell.number_format = fmt
            cell.border = BORDER_ALL
            if i in (3, 4, 5, 13, 20):  # 文字列セル: 左寄せ
                cell.alignment = LEFT
            elif i == 2 or i == 7:
                cell.alignment = CENTER
            else:
                cell.alignment = RIGHT
            if r % 2 == 0:
                cell.fill = FILL_ALT

    # === 自動計算式 ===
    for r in range(2, MAX_TRADES + 2):
        # A: # (連番、ただし B 列にデータが入っている時のみ)
        ws.cell(row=r, column=1, value=f'=IF(ISBLANK(B{r}),"",COUNTA($B$2:B{r}))')
        # N: SL pips = ABS(Entry - SL) / pip_size  (pip_size は通貨ペアから取得)
        ws.cell(row=r, column=14,
                value=f'=IFERROR(IF(ISBLANK(H{r}),"",ABS(H{r}-I{r})/VLOOKUP(C{r},設定!$B$9:$D$23,2,FALSE)),"")')
        # O: 計画 RR = ABS(TP - Entry) / ABS(Entry - SL)
        ws.cell(row=r, column=15,
                value=f'=IFERROR(IF(ISBLANK(J{r}),"",ABS(J{r}-H{r})/ABS(H{r}-I{r})),"")')
        # P: 損益 pips = IF(Long, (決済 - Entry), (Entry - 決済)) / pip_size
        ws.cell(row=r, column=16,
                value=f'=IFERROR(IF(OR(ISBLANK(L{r}),ISBLANK(E{r})),"",'
                      f'IF(E{r}="Long",(L{r}-H{r}),(H{r}-L{r}))/VLOOKUP(C{r},設定!$B$9:$D$23,2,FALSE)),"")')
        # Q: R-multiple = 損益 pips / SL pips
        ws.cell(row=r, column=17,
                value=f'=IFERROR(IF(OR(ISBLANK(P{r}),N{r}=0),"",P{r}/N{r}),"")')
        # R: 累計 R = 前行累計 + 今行 R
        if r == 2:
            ws.cell(row=r, column=18, value=f'=IFERROR(IF(ISBLANK(Q{r}),"",Q{r}),"")')
        else:
            ws.cell(row=r, column=18,
                    value=f'=IFERROR(IF(ISBLANK(Q{r}),R{r-1},IF(ISNUMBER(R{r-1}),R{r-1},0)+Q{r}),"")')
        # S: 損益 円 = pips × lot × pip_value_JPY
        ws.cell(row=r, column=19,
                value=f'=IFERROR(IF(OR(ISBLANK(P{r}),ISBLANK(K{r})),"",'
                      f'P{r}*K{r}*VLOOKUP(C{r},設定!$B$9:$D$23,3,FALSE)),"")')

    # === データ検証 (ドロップダウン) ===
    dv_pair   = DataValidation(type="list", formula1=f'"{",".join(PAIRS)}"', allow_blank=True)
    dv_method = DataValidation(type="list", formula1=f'"{",".join(METHODS)}"', allow_blank=True)
    dv_dir    = DataValidation(type="list", formula1=f'"{",".join(DIRECTIONS)}"', allow_blank=True)
    dv_reason = DataValidation(type="list", formula1=f'"{",".join(REASONS)}"', allow_blank=True)
    ws.add_data_validation(dv_pair)
    ws.add_data_validation(dv_method)
    ws.add_data_validation(dv_dir)
    ws.add_data_validation(dv_reason)
    dv_pair.add(f"C2:C{MAX_TRADES+1}")
    dv_method.add(f"D2:D{MAX_TRADES+1}")
    dv_dir.add(f"E2:E{MAX_TRADES+1}")
    dv_reason.add(f"M2:M{MAX_TRADES+1}")

    # === 条件付き書式: R-multiple 列 ===
    ws.conditional_formatting.add(
        f"Q2:Q{MAX_TRADES+1}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=CLR_GREEN_BG),
                   font=Font(color=CLR_GREEN, bold=True))
    )
    ws.conditional_formatting.add(
        f"Q2:Q{MAX_TRADES+1}",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=CLR_RED_BG),
                   font=Font(color=CLR_RED, bold=True))
    )
    # 損益 ¥ 列も同様
    ws.conditional_formatting.add(
        f"S2:S{MAX_TRADES+1}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   font=Font(color=CLR_GREEN, bold=True))
    )
    ws.conditional_formatting.add(
        f"S2:S{MAX_TRADES+1}",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(color=CLR_RED, bold=True))
    )
    # 決済理由: TP=緑、SL=赤
    ws.conditional_formatting.add(
        f"M2:M{MAX_TRADES+1}",
        FormulaRule(formula=[f'$M2="TP"'],
                    fill=PatternFill("solid", fgColor=CLR_GREEN_BG),
                    font=Font(color=CLR_GREEN, bold=True))
    )
    ws.conditional_formatting.add(
        f"M2:M{MAX_TRADES+1}",
        FormulaRule(formula=[f'$M2="SL"'],
                    fill=PatternFill("solid", fgColor=CLR_RED_BG),
                    font=Font(color=CLR_RED, bold=True))
    )

    # ペイン固定 (ヘッダー + #、日付 列まで)
    ws.freeze_panes = "C2"

    # === サンプルデータ 3 件 ===
    samples = [
        # 日付, ペア, 手法, 方向, スコア, Entry時刻, Entry, SL, TP, ロット, 決済価格, 理由, メモ
        ("2026/5/15", "USD/JPY", "TRIPLE", "Long",  85, "10:30", 157.20, 156.80, 158.40, 0.10, 158.40, "TP",  "TRIPLE 合意、東京時間。3R hit"),
        ("2026/5/16", "EUR/USD", "PDHL",   "Short", 78, "16:15", 1.0850, 1.0890, 1.0770, 0.10, 1.0890, "SL",  "PDL break fail, ロンドン時間"),
        ("2026/5/17", "GBP/JPY", "TRIPLE", "Long",  88, "21:00", 195.50, 194.80, 197.60, 0.05, 195.80, "手動", "NY 時間に手動利確 (急騰警戒)"),
    ]
    for i, (date, pair, method, direction, score, entry_time, entry, sl, tp, lot, close, reason, memo) in enumerate(samples, 2):
        ws.cell(row=i, column=2, value=date)
        ws.cell(row=i, column=3, value=pair)
        ws.cell(row=i, column=4, value=method)
        ws.cell(row=i, column=5, value=direction)
        ws.cell(row=i, column=6, value=score)
        ws.cell(row=i, column=7, value=entry_time)
        ws.cell(row=i, column=8, value=entry)
        ws.cell(row=i, column=9, value=sl)
        ws.cell(row=i, column=10, value=tp)
        ws.cell(row=i, column=11, value=lot)
        ws.cell(row=i, column=12, value=close)
        ws.cell(row=i, column=13, value=reason)
        ws.cell(row=i, column=20, value=memo)


# ============================================================
# シート 3: 設定 (定数 + pip 価値テーブル)
# ============================================================
def build_settings(ws):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20

    # タイトル
    ws.merge_cells("B2:D2")
    title = ws["B2"]
    title.value = "⚙️  設定 / Configuration"
    title.font = FONT_TITLE
    title.alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[2].height = 30

    # ===== 基本設定 (B4:C6) =====
    settings = [
        ("開始残高 (¥)",      1_000_000, "¥#,##0"),
        ("1 トレード リスク (%)", 0.5,    "0.0\\%"),
        ("USD/JPY 概算レート",  155,     "0.00"),
    ]
    for i, (label, value, fmt) in enumerate(settings):
        r = 3 + i
        ws.cell(row=r, column=2, value=label).font = FONT_BASE
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=2).fill = FILL_PANEL
        ws.cell(row=r, column=2).border = BORDER_ALL

        ws.cell(row=r, column=3, value=value)
        ws.cell(row=r, column=3).font = FONT_BASE_BOLD
        ws.cell(row=r, column=3).number_format = fmt
        ws.cell(row=r, column=3).alignment = RIGHT
        ws.cell(row=r, column=3).fill = FILL_GOLD
        ws.cell(row=r, column=3).border = BORDER_ALL

    # ===== 通貨ペア別 pip 価値テーブル =====
    section_header(ws, "B7:D7", "💱  通貨ペア別 pip 設定")

    # ヘッダー
    pip_headers = ["通貨ペア", "pip サイズ", "1pip × 1 lot (¥)"]
    for i, h in enumerate(pip_headers):
        c = ws.cell(row=8, column=2+i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER_ALL

    # JPY クロス: pip = 0.01、1 pip × 1 lot = 1000 円
    # 非 JPY: pip = 0.0001、1 pip × 1 lot ≒ 10 × USDJPY 円
    pip_table = []
    for p in PAIRS:
        if p.endswith("/JPY"):
            pip_table.append((p, 0.01, 1000))
        else:
            # 0.0001 × 100000 = 10 (quote ccy), USDJPY 換算で 10 × 155 = 1550 程度
            pip_table.append((p, 0.0001, "=10*$C$5"))

    for i, (pair, pip, val) in enumerate(pip_table):
        r = 9 + i
        ws.cell(row=r, column=2, value=pair).font = FONT_BASE_BOLD
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=3, value=pip)
        ws.cell(row=r, column=3).number_format = "0.0000"
        ws.cell(row=r, column=3).alignment = RIGHT
        ws.cell(row=r, column=4, value=val)
        ws.cell(row=r, column=4).number_format = "¥#,##0"
        ws.cell(row=r, column=4).alignment = RIGHT
        for col in (2, 3, 4):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER_ALL
            if i % 2 == 1:
                cell.fill = FILL_ALT

    # ガイド
    ws.merge_cells("B25:D28")
    g = ws["B25"]
    g.value = (
        "📌 黄色いセル (C3〜C5) を自分の運用に合わせて変更してください。\n"
        "・「開始残高」: 口座にいくらから始めるか\n"
        "・「1 トレード リスク」: 1 つのトレードで失っても良い % (推奨 0.3〜1.0%)\n"
        "・「USD/JPY 概算レート」: 非 JPY ペアの円換算用。月次で更新を推奨"
    )
    g.font = FONT_NOTE
    g.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
    g.fill = FILL_GOLD
    g.border = BORDER_BOX


# ============================================================
if __name__ == "__main__":
    import os
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "FX_Trade_Journal.xlsx")
    make_workbook(out)
