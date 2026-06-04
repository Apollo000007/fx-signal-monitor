"""手動トレード専用 Trade Journal を生成。

5 手法ごとに別シートを用意し、各手法の戦績を独立追跡できる構成。
計画 vs 実際、トレード評価、学びノートなど **手動トレード特有のカラム** を含む。

シート構成:
  1. 📊 ダッシュボード   — 5 手法を横並びで比較
  2. ORZ                — ORZ のみのトレード履歴 + 学び
  3. PDHL               — PDHL のみ
  4. BOTH               — ORZ + PDHL 合意
  5. CLAUDE             — Claude Confluence
  6. TRIPLE             — 3 手法合意 (★)
  7. 設定                — 開始残高 + pip 価値テーブル
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule

# ============================================================
# テーマ
# ============================================================
CLR_HEADER_BG    = "1F1B4E"
CLR_HEADER_FG    = "F5ECD7"
CLR_GOLD         = "E9C46A"
CLR_GOLD_BG      = "FFF7E0"
CLR_GREEN        = "16A34A"
CLR_GREEN_BG     = "DCFCE7"
CLR_RED          = "DC2626"
CLR_RED_BG       = "FEE2E2"
CLR_ROW_ALT_BG   = "F8F4FF"
CLR_PANEL_BG     = "F3F0FF"
CLR_BORDER       = "C7C0E0"

# 各手法ごとのテーマカラー (タブ識別用)
METHOD_THEMES = {
    "ORZ":    {"accent": "22D3EE", "bg": "E0F7FB", "title_fg": "0E7490"},   # シアン
    "PDHL":   {"accent": "F0A93B", "bg": "FFF4E0", "title_fg": "9A6203"},   # アンバー
    "BOTH":   {"accent": "A855F7", "bg": "F3E8FF", "title_fg": "7E22CE"},   # 紫
    "CLAUDE": {"accent": "4ADE80", "bg": "DCFCE7", "title_fg": "166534"},   # 緑
    "TRIPLE": {"accent": "E9C46A", "bg": "FFF7E0", "title_fg": "92400E"},   # 金
}

METHOD_DESC = {
    "ORZ":    "SMA 20/50/100 + 一目雲。4H で trend / range / unclear 分類、押し目買い・戻り売り・ブレイク・レンジ逆張りを 15M トリガーで実行。",
    "PDHL":   "前日高値 (PDH) / 前日安値 (PDL) のブレイク → リテスト → フラッグ → プライスアクション → トリガー の 5 段階確認型。ダマシ回避が核心。",
    "BOTH":   "ORZ と PDHL が同一方向で合意したときのみ発火。独立な 2 手法の確証で勝率向上を狙う。",
    "CLAUDE": "HTF バイアス / 4H モメンタム / 15M ATR 収縮 / 20EMA プルバック / RSI 再奪取 / Donchian ブレイクの 6 エッジ合流。",
    "TRIPLE": "ORZ + PDHL + Claude の 3 手法すべてが合意したときのみ発火。最も厳しく、Backtest で唯一の +EV を実証 (PF 3.42)。",
}

FONT_BASE      = Font(name="Noto Sans JP", size=10)
FONT_BASE_BOLD = Font(name="Noto Sans JP", size=10, bold=True)
FONT_HEADER    = Font(name="Noto Sans JP", size=10, bold=True, color=CLR_HEADER_FG)
FONT_TITLE     = Font(name="Noto Sans JP", size=20, bold=True, color=CLR_HEADER_BG)
FONT_SUB       = Font(name="Noto Sans JP", size=11, color="6B5B95")
FONT_KPI       = Font(name="Noto Sans JP", size=18, bold=True, color=CLR_HEADER_BG)
FONT_LABEL     = Font(name="Noto Sans JP", size=9, color="6B5B95")
FONT_NOTE      = Font(name="Noto Sans JP", size=9, italic=True, color="6B5B95")

THIN  = Side(style="thin", color=CLR_BORDER)
THICK = Side(style="medium", color=CLR_HEADER_BG)
BORDER_ALL  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOX  = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

FILL_HEADER = PatternFill("solid", fgColor=CLR_HEADER_BG)
FILL_GOLD   = PatternFill("solid", fgColor=CLR_GOLD_BG)
FILL_PANEL  = PatternFill("solid", fgColor=CLR_PANEL_BG)
FILL_ALT    = PatternFill("solid", fgColor=CLR_ROW_ALT_BG)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center", indent=1)
RIGHT  = Alignment(horizontal="right", vertical="center")

PAIRS = [
    "USD/JPY", "EUR/USD", "GBP/USD", "AUD/USD", "NZD/USD",
    "USD/CAD", "USD/CHF",
    "EUR/JPY", "GBP/JPY", "AUD/JPY", "NZD/JPY",
    "CAD/JPY", "CHF/JPY",
    "EUR/GBP",
]
DIRECTIONS = ["Long", "Short"]
REASONS    = ["TP", "SL", "手動", "強制", "建値撤退"]
RATINGS    = ["S", "A", "B", "C", "D"]
METHODS    = ["TRIPLE", "ORZ", "PDHL", "BOTH", "CLAUDE"]

MAX_TRADES = 200  # 各手法シート 200 行


def make_workbook(path: str):
    wb = Workbook()
    wb.remove(wb.active)

    # 順番に作成
    ws_dash = wb.create_sheet("📊ダッシュボード")
    for m in METHODS:
        wb.create_sheet(m)
    ws_cfg = wb.create_sheet("設定")

    build_settings(ws_cfg)
    for m in METHODS:
        build_strategy_sheet(wb[m], m)
    build_dashboard(ws_dash)

    for ws in wb:
        ws.sheet_view.showGridLines = False

    wb.save(path)
    print(f"[ok] saved: {path}")


# ============================================================
# 比較ダッシュボード
# ============================================================
def build_dashboard(ws):
    widths = [2, 18, 14, 14, 14, 14, 14, 4]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ===== タイトル =====
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "📊  手動トレード戦績 · 5 手法 比較ダッシュボード"
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:G3")
    sub = ws["B3"]
    sub.value = "各手法のシートにトレードを記録 → ここで横並び比較。「どの手法が勝てるか」を客観評価。"
    sub.font = FONT_SUB
    sub.alignment = Alignment(horizontal="left", indent=1)

    # ===== 5 手法 横並びテーブル =====
    section_header(ws, "B5:G5", "🎯  手法別パフォーマンス (5 手法横並び)")

    # ヘッダー行 (行 6)
    ws.cell(row=6, column=2, value="指標").font = FONT_HEADER
    ws.cell(row=6, column=2).fill = FILL_HEADER
    ws.cell(row=6, column=2).alignment = LEFT
    ws.cell(row=6, column=2).border = BORDER_ALL

    for i, m in enumerate(METHODS):
        theme = METHOD_THEMES[m]
        cell = ws.cell(row=6, column=3+i, value=m)
        cell.font = Font(name="Noto Sans JP", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=theme["accent"])
        cell.alignment = CENTER
        cell.border = BORDER_ALL

    # 指標行 (各手法シートの集計セルを参照)
    metrics = [
        ("件数",           "TotalTrades",     "0"),
        ("勝率",           "WinRate",         "0.0%"),
        ("平均 RR (計画)", "AvgPlanRR",       "0.00"),
        ("PF",             "PF",              "0.00"),
        ("期待値 (R)",     "Expectancy",      "+0.000;-0.000;-"),
        ("累計 R",         "TotalR",          "+0.00;-0.00;-"),
        ("累計損益 ¥",     "TotalYen",        "¥#,##0;[Red]-¥#,##0;-"),
        ("最大 DD (R)",    "MaxDD",           "+0.00;-0.00;-"),
        ("勝ち平均 R",     "AvgWinR",         "+0.000;-0.000;-"),
        ("負け平均 R",     "AvgLossR",        "+0.000;-0.000;-"),
        ("評価 S 件数",    "SRated",          "0"),
        ("評価 D 件数",    "DRated",          "0"),
    ]

    for i, (label, key, fmt) in enumerate(metrics):
        r = 7 + i
        # ラベル
        ws.cell(row=r, column=2, value=label).font = FONT_BASE
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=2).fill = FILL_PANEL
        ws.cell(row=r, column=2).border = BORDER_ALL
        # 値
        for j, m in enumerate(METHODS):
            col = 3 + j
            # 各手法シートの A 列 (ラベル) と B 列 (値) を参照する想定
            # → 各手法シートに集計値を A2-A30 / B2-B30 に置く構造で
            #   ここでは式を「シート名!固定セル」で参照する
            #   集計セルの場所は build_strategy_sheet と整合
            cell_ref = SUMMARY_CELLS[key]  # 例: "B4"
            ws.cell(row=r, column=col, value=f"='{m}'!{cell_ref}").number_format = fmt
            ws.cell(row=r, column=col).font = FONT_BASE_BOLD
            ws.cell(row=r, column=col).alignment = CENTER
            ws.cell(row=r, column=col).border = BORDER_ALL
            if i % 2 == 1:
                ws.cell(row=r, column=col).fill = FILL_ALT

    # 累計 R 行に色付き条件
    total_r_row = 7 + 5  # "累計 R" の行
    ws.conditional_formatting.add(
        f"C{total_r_row}:G{total_r_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=CLR_GREEN_BG),
                   font=Font(color=CLR_GREEN, bold=True))
    )
    ws.conditional_formatting.add(
        f"C{total_r_row}:G{total_r_row}",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=CLR_RED_BG),
                   font=Font(color=CLR_RED, bold=True))
    )
    # 期待値 (R) 行も
    ev_row = 7 + 4
    for row_to_color in (ev_row, total_r_row):
        ws.conditional_formatting.add(
            f"C{row_to_color}:G{row_to_color}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=PatternFill("solid", fgColor=CLR_GREEN_BG))
        )
        ws.conditional_formatting.add(
            f"C{row_to_color}:G{row_to_color}",
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=PatternFill("solid", fgColor=CLR_RED_BG))
        )

    # ===== 集計サマリ =====
    summary_row = 7 + len(metrics) + 2  # 最後のメトリクス行 + 2
    section_header(ws, f"B{summary_row}:G{summary_row}", "💡  全体サマリ")

    summary_items = [
        ("総トレード数 (全手法合計)",
         "=" + "+".join([f"'{m}'!{SUMMARY_CELLS['TotalTrades']}" for m in METHODS]),
         "0"),
        ("全手法 合計累計 R",
         "=" + "+".join([f"'{m}'!{SUMMARY_CELLS['TotalR']}" for m in METHODS]),
         "+0.00;-0.00;-"),
        ("全手法 合計累計損益 ¥",
         "=" + "+".join([f"'{m}'!{SUMMARY_CELLS['TotalYen']}" for m in METHODS]),
         "¥#,##0;[Red]-¥#,##0;-"),
        ("現在残高",
         "=設定!C3+(" + "+".join([f"'{m}'!{SUMMARY_CELLS['TotalYen']}" for m in METHODS]) + ")",
         "¥#,##0"),
    ]
    for i, (label, formula, fmt) in enumerate(summary_items):
        r = summary_row + 1 + i
        ws.cell(row=r, column=2, value=label).font = FONT_BASE
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=2).fill = FILL_PANEL
        ws.cell(row=r, column=2).border = BORDER_ALL
        # 値 (C:G にマージ)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.cell(row=r, column=3, value=formula)
        ws.cell(row=r, column=3).font = FONT_BASE_BOLD
        ws.cell(row=r, column=3).alignment = RIGHT
        ws.cell(row=r, column=3).number_format = fmt
        ws.cell(row=r, column=3).fill = FILL_GOLD
        for col in range(3, 8):
            ws.cell(row=r, column=col).border = BORDER_ALL

    # ===== ガイド =====
    guide_row = summary_row + len(summary_items) + 3
    ws.merge_cells(start_row=guide_row, start_column=2, end_row=guide_row+3, end_column=7)
    cell = ws.cell(row=guide_row, column=2)
    cell.value = (
        "💡 使い方:\n"
        "1. 手法シート (ORZ / PDHL / BOTH / CLAUDE / TRIPLE) でトレードを記録\n"
        "2. このダッシュボードで横並び比較\n"
        "3. 累計 R が +EV の手法を継続、 -EV を停止判断\n"
        "4. 「評価」列で S/A ランクが多い手法 = 自分のスキルが乗っている手法"
    )
    cell.font = FONT_NOTE
    cell.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
    cell.fill = FILL_GOLD
    cell.border = BORDER_BOX
    ws.row_dimensions[guide_row].height = 22
    for r in range(guide_row, guide_row + 4):
        ws.row_dimensions[r].height = 22


# 集計セルのマッピング (各手法シートの固定セル)
SUMMARY_CELLS = {
    "TotalTrades": "C4",
    "WinRate":     "C5",
    "AvgPlanRR":   "C6",
    "PF":          "C7",
    "Expectancy":  "C8",
    "TotalR":      "C9",
    "TotalYen":    "C10",
    "MaxDD":       "C11",
    "AvgWinR":     "C12",
    "AvgLossR":    "C13",
    "SRated":      "C14",
    "DRated":      "C15",
}


# ============================================================
# 各手法シート
# ============================================================
def build_strategy_sheet(ws, method: str):
    theme = METHOD_THEMES[method]
    description = METHOD_DESC[method]

    # --- 列幅 ---
    widths = {
        1: 5,   # A 番号
        2: 12,  # B 日付
        3: 11,  # C 通貨ペア
        4: 7,   # D 方向
        5: 7,   # E スコア
        6: 11,  # F 計画Entry
        7: 11,  # G 実Entry
        8: 11,  # H 計画SL
        9: 11,  # I 実SL
        10: 11, # J 計画TP
        11: 11, # K 実TP
        12: 7,  # L ロット
        13: 11, # M 決済価格
        14: 10, # N 決済理由
        15: 9,  # O SL pips
        16: 8,  # P 計画RR
        17: 10, # Q 損益 pips
        18: 9,  # R R-mult
        19: 9,  # S 累計R
        20: 12, # T 損益¥
        21: 8,  # U 評価
        22: 35, # V 学び
    }
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w

    # ===== トップ: タイトル =====
    ws.merge_cells("A1:V1")
    title = ws["A1"]
    title.value = f"{method}  ·  手動トレード履歴"
    title.font = Font(name="Noto Sans JP", size=20, bold=True, color=theme["title_fg"])
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    title.fill = PatternFill("solid", fgColor=theme["bg"])
    ws.row_dimensions[1].height = 36

    # ===== 手法説明 =====
    ws.merge_cells("A2:V2")
    desc = ws["A2"]
    desc.value = description
    desc.font = FONT_SUB
    desc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    desc.fill = PatternFill("solid", fgColor=theme["bg"])
    ws.row_dimensions[2].height = 32

    # ===== 集計セル (ダッシュボード用) — 行 3-15 列 B-C =====
    # B 列: ラベル、 C 列: 値 (formula)
    summary = [
        ("📊 集計サマリ",          None,           None),                                    # 3
        ("件数",                   "TotalTrades",  '=COUNT(R18:R{end})'),                    # 4
        ("勝率",                   "WinRate",      '=IFERROR(COUNTIF(R18:R{end},">0")/COUNT(R18:R{end}),0)'),  # 5
        ("平均 RR (計画)",         "AvgPlanRR",    '=IFERROR(AVERAGE(P18:P{end}),0)'),       # 6
        ("PF",                     "PF",           '=IFERROR(SUMIF(R18:R{end},">0")/-SUMIF(R18:R{end},"<0"),0)'),  # 7
        ("期待値 (R)",             "Expectancy",   '=IFERROR(AVERAGE(R18:R{end}),0)'),       # 8
        ("累計 R",                 "TotalR",       '=SUM(R18:R{end})'),                      # 9
        ("累計損益 ¥",             "TotalYen",     '=SUM(T18:T{end})'),                      # 10
        ("最大 DD (R)",            "MaxDD",        '=IFERROR(MAX(S18:S{end})-MIN(S18:S{end}),0)'),  # 11
        ("勝ち平均 R",             "AvgWinR",      '=IFERROR(AVERAGEIF(R18:R{end},">0"),0)'),  # 12
        ("負け平均 R",             "AvgLossR",     '=IFERROR(AVERAGEIF(R18:R{end},"<0"),0)'),  # 13
        ("評価 S 件数",            "SRated",       f'=COUNTIF(U18:U{{end}},"S")'),           # 14
        ("評価 D 件数",            "DRated",       f'=COUNTIF(U18:U{{end}},"D")'),           # 15
    ]

    fmts = {
        "TotalTrades": "0", "WinRate": "0.0%", "AvgPlanRR": "0.00",
        "PF": "0.00", "Expectancy": "+0.000;-0.000;-", "TotalR": "+0.00;-0.00;-",
        "TotalYen": "¥#,##0;[Red]-¥#,##0;-", "MaxDD": "+0.00;-0.00;-",
        "AvgWinR": "+0.000;-0.000;-", "AvgLossR": "+0.000;-0.000;-",
        "SRated": "0", "DRated": "0",
    }

    end_row = MAX_TRADES + 17  # トレード行は 18 から始まる

    for i, (label, key, formula) in enumerate(summary):
        r = 3 + i
        # ラベル (B 列)
        cell_b = ws.cell(row=r, column=2, value=label)
        if i == 0:  # ヘッダー行
            cell_b.font = FONT_HEADER
            cell_b.fill = FILL_HEADER
            cell_b.alignment = LEFT
            cell_b.border = BORDER_ALL
            # C も同じ
            cell_c = ws.cell(row=r, column=3, value="")
            cell_c.fill = FILL_HEADER
            cell_c.border = BORDER_ALL
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            continue

        cell_b.font = FONT_BASE
        cell_b.alignment = LEFT
        cell_b.fill = FILL_PANEL
        cell_b.border = BORDER_ALL
        # 値 (C 列)
        formula_filled = formula.format(end=end_row)
        cell_c = ws.cell(row=r, column=3, value=formula_filled)
        cell_c.font = FONT_BASE_BOLD
        cell_c.alignment = RIGHT
        cell_c.number_format = fmts[key]
        cell_c.fill = FILL_GOLD if key in ("TotalR", "TotalYen", "Expectancy") else FILL_PANEL
        cell_c.border = BORDER_ALL

    # ===== 評価分布バー (D3-G15 を使ってカウント表示) =====
    ws.merge_cells("E3:H3")
    sec = ws.cell(row=3, column=5, value="🏅 トレード評価分布")
    sec.font = FONT_HEADER
    sec.fill = FILL_HEADER
    sec.alignment = LEFT
    sec.border = BORDER_ALL

    rating_colors = {"S": "FFD700", "A": "90EE90", "B": "87CEEB", "C": "FFA07A", "D": "FF6B6B"}
    for i, rating in enumerate(RATINGS):
        r = 4 + i
        ws.cell(row=r, column=5, value=f"{rating} ランク").font = FONT_BASE
        ws.cell(row=r, column=5).alignment = LEFT
        ws.cell(row=r, column=5).fill = FILL_PANEL
        ws.cell(row=r, column=5).border = BORDER_ALL

        ws.cell(row=r, column=6, value=f'=COUNTIF(U18:U{end_row},"{rating}")').number_format = "0"
        ws.cell(row=r, column=6).font = FONT_BASE_BOLD
        ws.cell(row=r, column=6).alignment = CENTER
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=rating_colors[rating])
        ws.cell(row=r, column=6).border = BORDER_ALL

        # 割合
        ws.cell(row=r, column=7, value=f'=IFERROR(F{r}/COUNTA(U18:U{end_row}),0)').number_format = "0.0%"
        ws.cell(row=r, column=7).font = FONT_BASE
        ws.cell(row=r, column=7).alignment = CENTER
        ws.cell(row=r, column=7).fill = FILL_PANEL
        ws.cell(row=r, column=7).border = BORDER_ALL

        ws.cell(row=r, column=8, value=_rating_meaning(rating)).font = FONT_NOTE
        ws.cell(row=r, column=8).alignment = LEFT
        ws.cell(row=r, column=8).fill = FILL_PANEL
        ws.cell(row=r, column=8).border = BORDER_ALL

    # ===== 学びメモ (I3-V15) =====
    ws.merge_cells("I3:V3")
    sec2 = ws.cell(row=3, column=9, value=f"📝 {method} で学んだこと · 観察メモ")
    sec2.font = FONT_HEADER
    sec2.fill = FILL_HEADER
    sec2.alignment = LEFT
    sec2.border = BORDER_ALL

    ws.merge_cells("I4:V15")
    notes = ws.cell(row=4, column=9)
    notes.value = (
        f"(ここに {method} で気づいたこと・パターン・反省を自由記述。例:\n"
        " ・東京時間の発火は勝率が低い、ロンドン時間に絞ると改善\n"
        " ・スコア 80+ のシグナルだけ取ると勝率が劇的に上がる\n"
        " ・SL を少しタイトめにすると RR 向上)"
    )
    notes.font = Font(name="Noto Sans JP", size=10, color="6B5B95", italic=True)
    notes.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    notes.fill = PatternFill("solid", fgColor=theme["bg"])
    notes.border = BORDER_BOX

    # ===== トレード履歴ヘッダー (行 17) =====
    headers = [
        "#", "日付", "通貨ペア", "方向", "スコア",
        "計画Entry", "実Entry", "計画SL", "実SL", "計画TP", "実TP",
        "ロット", "決済価格", "決済理由",
        "実SL pips", "計画RR", "損益 pips", "R-mult", "累計R", "損益 ¥",
        "評価", "学び / 反省"
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=17, column=i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[17].height = 28

    # ===== トレード行 (18 〜 17+MAX) =====
    fmts_col = {
        1: "0", 2: "yyyy/mm/dd", 3: "@", 4: "@", 5: "0",
        6: "0.00000", 7: "0.00000", 8: "0.00000", 9: "0.00000",
        10: "0.00000", 11: "0.00000", 12: "0.00",
        13: "0.00000", 14: "@",
        15: "0.0", 16: "0.00", 17: "+0.0;-0.0;-",
        18: "+0.00;-0.00;-", 19: "+0.00;-0.00;-",
        20: "¥#,##0;[Red]-¥#,##0;-",
        21: "@", 22: "@",
    }

    for r in range(18, end_row + 1):
        for c in range(1, 23):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_BASE
            cell.number_format = fmts_col[c]
            cell.border = BORDER_ALL
            if c in (3, 4, 14, 21, 22):
                cell.alignment = LEFT
            elif c == 2:
                cell.alignment = CENTER
            else:
                cell.alignment = RIGHT
            if r % 2 == 1:
                cell.fill = FILL_ALT

        # --- 自動計算式 ---
        # A: 番号
        ws.cell(row=r, column=1, value=f'=IF(ISBLANK(B{r}),"",COUNTA($B$18:B{r}))')
        # O: 実SL pips = ABS(実Entry - 実SL) / pip_size
        ws.cell(row=r, column=15,
                value=f'=IFERROR(IF(OR(ISBLANK(G{r}),ISBLANK(I{r})),"",'
                      f'ABS(G{r}-I{r})/VLOOKUP(C{r},設定!$B$9:$D$23,2,FALSE)),"")')
        # P: 計画RR = ABS(計画TP - 計画Entry) / ABS(計画Entry - 計画SL)
        ws.cell(row=r, column=16,
                value=f'=IFERROR(IF(OR(ISBLANK(F{r}),ISBLANK(H{r}),ISBLANK(J{r})),"",'
                      f'ABS(J{r}-F{r})/ABS(F{r}-H{r})),"")')
        # Q: 損益 pips
        ws.cell(row=r, column=17,
                value=f'=IFERROR(IF(OR(ISBLANK(M{r}),ISBLANK(D{r}),ISBLANK(G{r})),"",'
                      f'IF(D{r}="Long",(M{r}-G{r}),(G{r}-M{r}))/VLOOKUP(C{r},設定!$B$9:$D$23,2,FALSE)),"")')
        # R: R-multiple = 損益 pips / SL pips
        ws.cell(row=r, column=18,
                value=f'=IFERROR(IF(OR(ISBLANK(Q{r}),O{r}=0),"",Q{r}/O{r}),"")')
        # S: 累計 R
        if r == 18:
            ws.cell(row=r, column=19, value=f'=IFERROR(IF(ISBLANK(R{r}),"",R{r}),"")')
        else:
            ws.cell(row=r, column=19,
                    value=f'=IFERROR(IF(ISBLANK(R{r}),S{r-1},IF(ISNUMBER(S{r-1}),S{r-1},0)+R{r}),"")')
        # T: 損益 ¥
        ws.cell(row=r, column=20,
                value=f'=IFERROR(IF(OR(ISBLANK(Q{r}),ISBLANK(L{r})),"",'
                      f'Q{r}*L{r}*VLOOKUP(C{r},設定!$B$9:$D$23,3,FALSE)),"")')

    # === データ検証 (ドロップダウン) ===
    dv_pair   = DataValidation(type="list", formula1=f'"{",".join(PAIRS)}"', allow_blank=True)
    dv_dir    = DataValidation(type="list", formula1=f'"{",".join(DIRECTIONS)}"', allow_blank=True)
    dv_reason = DataValidation(type="list", formula1=f'"{",".join(REASONS)}"', allow_blank=True)
    dv_rating = DataValidation(type="list", formula1=f'"{",".join(RATINGS)}"', allow_blank=True)
    ws.add_data_validation(dv_pair)
    ws.add_data_validation(dv_dir)
    ws.add_data_validation(dv_reason)
    ws.add_data_validation(dv_rating)
    dv_pair.add(f"C18:C{end_row}")
    dv_dir.add(f"D18:D{end_row}")
    dv_reason.add(f"N18:N{end_row}")
    dv_rating.add(f"U18:U{end_row}")

    # === 条件付き書式 ===
    # R-multiple: 緑/赤
    ws.conditional_formatting.add(
        f"R18:R{end_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=CLR_GREEN_BG),
                   font=Font(color=CLR_GREEN, bold=True))
    )
    ws.conditional_formatting.add(
        f"R18:R{end_row}",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=CLR_RED_BG),
                   font=Font(color=CLR_RED, bold=True))
    )
    # 決済理由: TP=緑、SL=赤
    ws.conditional_formatting.add(
        f"N18:N{end_row}",
        FormulaRule(formula=['$N18="TP"'],
                    fill=PatternFill("solid", fgColor=CLR_GREEN_BG),
                    font=Font(color=CLR_GREEN, bold=True))
    )
    ws.conditional_formatting.add(
        f"N18:N{end_row}",
        FormulaRule(formula=['$N18="SL"'],
                    fill=PatternFill("solid", fgColor=CLR_RED_BG),
                    font=Font(color=CLR_RED, bold=True))
    )
    # 評価ランク: S=金、A=緑、D=赤
    rating_fills = {
        "S": ("FFF7E0", "B8860B"),
        "A": ("DCFCE7", CLR_GREEN),
        "D": ("FEE2E2", CLR_RED),
    }
    for rating, (bg, fg) in rating_fills.items():
        ws.conditional_formatting.add(
            f"U18:U{end_row}",
            FormulaRule(formula=[f'$U18="{rating}"'],
                        fill=PatternFill("solid", fgColor=bg),
                        font=Font(color=fg, bold=True))
        )

    # ペイン固定: ヘッダー (17行目まで)
    ws.freeze_panes = "C18"


def _rating_meaning(rating):
    return {
        "S": "完璧 (計画通り + 理想的展開)",
        "A": "良い (利益確保)",
        "B": "普通 (微益微損)",
        "C": "悪い (損失だが学びあり)",
        "D": "失敗 (規律違反含む)",
    }.get(rating, "")


# ============================================================
# 設定シート
# ============================================================
def build_settings(ws):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20

    # タイトル
    ws.merge_cells("B2:D2")
    title = ws["B2"]
    title.value = "⚙️  設定 / Configuration"
    title.font = FONT_TITLE
    title.alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[2].height = 30

    # 基本設定
    settings = [
        ("開始残高 (¥)",      1_000_000, "¥#,##0"),
        ("1 トレード リスク (%)", 0.5,    "0.0\\%"),
        ("USD/JPY 概算レート",  155,     "0.00"),
    ]
    for i, (label, value, fmt) in enumerate(settings):
        r = 3 + i
        ws.cell(row=r, column=2, value=label).font = FONT_BASE
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=2).fill = FILL_PANEL
        ws.cell(row=r, column=2).border = BORDER_ALL

        ws.cell(row=r, column=3, value=value)
        ws.cell(row=r, column=3).font = FONT_BASE_BOLD
        ws.cell(row=r, column=3).number_format = fmt
        ws.cell(row=r, column=3).alignment = RIGHT
        ws.cell(row=r, column=3).fill = FILL_GOLD
        ws.cell(row=r, column=3).border = BORDER_ALL

    # pip テーブル
    section_header(ws, "B7:D7", "💱  通貨ペア別 pip 設定")
    pip_headers = ["通貨ペア", "pip サイズ", "1pip × 1 lot (¥)"]
    for i, h in enumerate(pip_headers):
        c = ws.cell(row=8, column=2+i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER_ALL

    pip_table = []
    for p in PAIRS:
        if p.endswith("/JPY"):
            pip_table.append((p, 0.01, 1000))
        else:
            pip_table.append((p, 0.0001, "=10*$C$5"))

    for i, (pair, pip, val) in enumerate(pip_table):
        r = 9 + i
        ws.cell(row=r, column=2, value=pair).font = FONT_BASE_BOLD
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=3, value=pip)
        ws.cell(row=r, column=3).number_format = "0.0000"
        ws.cell(row=r, column=3).alignment = RIGHT
        ws.cell(row=r, column=4, value=val)
        ws.cell(row=r, column=4).number_format = "¥#,##0"
        ws.cell(row=r, column=4).alignment = RIGHT
        for col in (2, 3, 4):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER_ALL
            if i % 2 == 1:
                cell.fill = FILL_ALT

    # ガイド
    ws.merge_cells("B25:D30")
    g = ws["B25"]
    g.value = (
        "📌 このスプレッドシートは「手動トレード」専用です。\n\n"
        "5 手法のシート (ORZ/PDHL/BOTH/CLAUDE/TRIPLE) に\n"
        "それぞれのトレードを記録すると、📊ダッシュボードで\n"
        "横並び比較ができます。\n\n"
        "黄色いセル (C3〜C5) は自分の運用に合わせて変更可。"
    )
    g.font = FONT_NOTE
    g.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
    g.fill = FILL_GOLD
    g.border = BORDER_BOX


def section_header(ws, range_str, text):
    ws.merge_cells(range_str)
    first = range_str.split(":")[0]
    cell = ws[first]
    cell.value = text
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    row = int("".join(filter(str.isdigit, first)))
    ws.row_dimensions[row].height = 22


# ============================================================
if __name__ == "__main__":
    import os
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "FX_Manual_Trade_Journal.xlsx")
    make_workbook(out)
